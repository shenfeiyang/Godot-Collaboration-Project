import openpyxl

wb = openpyxl.load_workbook("execl/equip.xlsx", data_only=False)
ws = wb["equip"]

# Find header & skillSp column
header_row = None
skill_sp_col = None
for i, row in enumerate(ws.iter_rows(min_row=1, max_row=20, values_only=True), 1):
    if row and row[0] == "id":
        header_row = i
        for j, val in enumerate(row):
            if val and str(val).strip() == "skillSp":
                skill_sp_col = j + 1  # openpyxl is 1-indexed
        break

print(f"Header at row {header_row}, skillSp at col {skill_sp_col}")

# Map: equipment name → skill ID
SKILL_MAP = {
    "手枪01": "skill_002",
    "机关枪01": "skill_003",
    "激光01": "skill_007",
    "喷火器01": "skill_005",
}

if header_row and skill_sp_col:
    for row in ws.iter_rows(min_row=header_row + 2, max_row=ws.max_row):
        name_cell = row[1]  # name is column 2 (0-indexed = 1)
        skill_cell = row[skill_sp_col - 1]  # adjust for 0-index
        if name_cell.value and str(name_cell.value).strip() in SKILL_MAP:
            old_val = str(skill_cell.value or "")
            new_val = SKILL_MAP[str(name_cell.value).strip()]
            if old_val != new_val:
                print(f"  {name_cell.value:12s} skillSp: '{old_val}' → '{new_val}'")
                skill_cell.value = new_val

wb.save("execl/equip.xlsx")
print("Done!")
